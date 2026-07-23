import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(bold=True)


def write_sheet(ws, title, rows):
    ws.title = title
    ws.append(["계정과목", "전기", "당기"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16


bs_rows = [
    ["현금및현금성자산", 0, 0],
    ["매출채권", 0, 0],
    ["재고자산", 0, 0],
    ["기타유동자산", 0, 0],
    ["유동자산합계", 0, 0],
    ["유형자산", 0, 0],
    ["무형자산", 0, 0],
    ["투자자산", 0, 0],
    ["기타비유동자산", 0, 0],
    ["비유동자산합계", 0, 0],
    ["자산총계", 0, 0],
    ["매입채무", 0, 0],
    ["단기차입금", 0, 0],
    ["기타유동부채", 0, 0],
    ["유동부채합계", 0, 0],
    ["장기차입금", 0, 0],
    ["기타비유동부채", 0, 0],
    ["비유동부채합계", 0, 0],
    ["부채총계", 0, 0],
    ["자본금", 0, 0],
    ["자본잉여금", 0, 0],
    ["이익잉여금", 0, 0],
    ["기타자본", 0, 0],
    ["자본총계", 0, 0],
]

is_rows = [
    ["매출액", 0, 0],
    ["매출원가", 0, 0],
    ["매출총이익", 0, 0],
    ["판매비와관리비", 0, 0],
    ["영업이익", 0, 0],
    ["영업외수익", 0, 0],
    ["영업외비용", 0, 0],
    ["법인세비용차감전순이익", 0, 0],
    ["법인세비용", 0, 0],
    ["당기순이익", 0, 0],
    ["감가상각비", 0, 0],
]

cf_rows = [
    ["영업활동현금흐름", 0, 0],
    ["투자활동현금흐름", 0, 0],
    ["재무활동현금흐름", 0, 0],
    ["현금및현금성자산의순증감", 0, 0],
    ["기초현금", 0, 0],
    ["기말현금", 0, 0],
]

je_ws = wb.active
write_sheet(je_ws, "재무상태표", bs_rows)

is_ws = wb.create_sheet()
write_sheet(is_ws, "손익계산서", is_rows)

cf_ws = wb.create_sheet()
write_sheet(cf_ws, "현금흐름표", cf_rows)

journal_ws = wb.create_sheet()
journal_ws.title = "전표데이터"
journal_headers = ["전표번호", "전기일자", "전기시각", "계정과목", "차변", "대변", "작성자", "승인자", "적요"]
journal_ws.append(journal_headers)
for cell in journal_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
for col, width in zip("ABCDEFGHI", [12, 14, 10, 18, 14, 14, 12, 12, 24]):
    journal_ws.column_dimensions[col].width = width

wb.save("templates/financial_template.xlsx")
print("템플릿 생성 완료: templates/financial_template.xlsx")
